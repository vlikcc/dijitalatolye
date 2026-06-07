#!/usr/bin/env python3
"""
MEB kazanım Excel dosyalarını Catalog import JSON'una dönüştürür.

Kaynak:  data/meb-kazanimlar/*.xlsx  (her dosya = ders, her sayfa genelde = sınıf)
Çıktı:   src/Services/Catalog/Catalog.API/Seed/meb-outcomes.json
         [{ gradeCode, subjectCode, subjectName, unitName, unitOrder,
            outcomeCode, outcomeDescription }, ...]  (outcomeCode'a göre dedupe)

3 layout başlık satırından ayırt edilir:
  - standart:   Kazanım Kodu | Ünite/Tema No | Ünite/Tema Adı | Öğrenme Çıktısı   (sayfa=sınıf)
  - tek-sayfa:  Sınıf Düzeyi | Kazanım Kodu | Ünite/Tema No & Adı | Öğrenme Çıktısı (sınıf kolonda)
  - okul öncesi: Kazanım Kodu | Alt Alan/Alt Beceri | Öğrenme Çıktısı            (sayfa=alt alan)

Kullanım:
  python3 convert_kazanimlar.py [--src DIR] [--out FILE]
"""
import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl gerekli: pip install openpyxl")

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DEFAULT_SRC = os.path.join(os.path.dirname(__file__), "xlsx")
DEFAULT_OUT = os.path.join(REPO, "src", "Services", "Catalog", "Catalog.API", "Seed", "meb-outcomes.json")

# Dosya adı (lower) substring -> (kanonik ders adı, ders kodu).
# MEB başlıkları büyük harf + tutarsız noktalı/noktasız I içerdiğinden, ders adını dosyadan kanonikleştiriyoruz.
# İki Din ve iki İnkılap dosyası aynı koda map'lenir (birleşir). Kodlar mevcut seed'le hizalı (matematik/turkce/fen/sosyal/ingilizce).
FILENAME_SUBJECTS = [
    ("bilişim", ("Bilişim Teknolojileri", "bilisim")),
    ("bilisim", ("Bilişim Teknolojileri", "bilisim")),
    ("biyoloji", ("Biyoloji", "biyoloji")),
    ("bilsem", ("BİLSEM", "bilsem")),
    ("coğrafya", ("Coğrafya", "cografya")),
    ("cografya", ("Coğrafya", "cografya")),
    ("din", ("Din Kültürü ve Ahlak Bilgisi", "din")),
    ("fen", ("Fen Bilimleri", "fen")),
    ("fizik", ("Fizik", "fizik")),
    ("hayat", ("Hayat Bilgisi", "hayat-bilgisi")),
    ("kimya", ("Kimya", "kimya")),
    ("matematik", ("Matematik", "matematik")),
    ("öncesi", ("Okul Öncesi", "okul-oncesi")),
    ("oncesi", ("Okul Öncesi", "okul-oncesi")),
    ("sosyal", ("Sosyal Bilgiler", "sosyal")),
    ("türkdili", ("Türk Dili ve Edebiyatı", "turk-dili-edebiyat")),
    ("türkçe", ("Türkçe", "turkce")),
    ("turkce", ("Türkçe", "turkce")),
    ("ingilizce", ("İngilizce", "ingilizce")),
    ("inkılap", ("T.C. İnkılap Tarihi ve Atatürkçülük", "inkilap")),
    ("inkilap", ("T.C. İnkılap Tarihi ve Atatürkçülük", "inkilap")),
    ("tarih", ("Tarih", "tarih")),  # "tarih" en sonda: inkılap/türkdili önce eşleşsin
]


def resolve_subject(filename):
    low = os.path.basename(filename).translate(TR_LOWER).lower()
    for needle, (name, code) in FILENAME_SUBJECTS:
        if needle.translate(TR_LOWER).lower() in low:
            return name, code
    base = re.sub(r"\.xlsx$", "", os.path.basename(filename), flags=re.IGNORECASE)
    return base.strip(), slugify(base)

TR_LOWER = str.maketrans("İIÇĞÖŞÜ", "iıçğöşü")


def slugify(s: str) -> str:
    s = (s or "").translate(TR_LOWER).lower()
    s = (s.replace("ç", "c").replace("ğ", "g").replace("ı", "i")
          .replace("ö", "o").replace("ş", "s").replace("ü", "u"))
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "ders"


def subject_code(name: str) -> str:
    key = slugify(name).replace("-", " ")
    return SUBJECT_CODE_ALIASES.get(key, slugify(name))


def parse_subject_name(title, filename):
    """Satır 1 başlığından ders adını çıkar; yoksa dosya adından."""
    if title:
        t = str(title)
        m = re.search(r"—\s*(.+?)\s+(?:DERS[İI]|ÖĞRET[İI]M|PROGRAMI)", t, re.IGNORECASE)
        if m:
            return _titlecase_tr(m.group(1).strip())
    base = re.sub(r"\.xlsx$", "", os.path.basename(filename), flags=re.IGNORECASE)
    base = re.sub(r"\s*\d+\s*-\s*\d+.*$", "", base)         # "1-12. sınıf" vb. at
    base = re.sub(r"\s*\d+\.?\s*s[ıi]n[ıi]f.*$", "", base, flags=re.IGNORECASE)
    base = base.replace("SON", "").strip()
    return _titlecase_tr(base or "Ders")


def _titlecase_tr(s: str) -> str:
    words = []
    for w in s.split():
        if not w:
            continue
        if w.isupper() and len(w) > 1:
            first = w[0]
            rest = w[1:].translate(TR_LOWER).lower()
            words.append(first + rest)
        else:
            words.append(w)
    return " ".join(words)


def grade_from_sheet(name: str):
    n = str(name).strip()
    m = re.search(r"(\d+)", n)
    if m:
        return m.group(1)
    low = n.translate(TR_LOWER).lower()
    if "hazırl" in low or "hazirl" in low:
        return "hazirlik"
    return None  # bilinmiyor -> arayan karar versin


def first_int(s):
    if s is None:
        return None
    m = re.search(r"(\d+)", str(s))
    return int(m.group(1)) if m else None


def find_header(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, max_col=8, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        joined = " | ".join(cells)
        if "Kazanım Kodu" in joined:
            return i + 1, cells
    return None, None


def col_index(header, *needles):
    for idx, h in enumerate(header):
        hl = h.lower()
        if all(n.lower() in hl for n in needles):
            return idx
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    files = sorted(f for f in os.listdir(args.src) if f.lower().endswith(".xlsx") and not f.startswith("~"))
    if not files:
        sys.exit(f"Excel bulunamadı: {args.src}")

    rows = []
    seen_codes = set()
    skipped = 0

    for fname in files:
        path = os.path.join(args.src, fname)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        subj_name, subj_code = resolve_subject(fname)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            hdr_row, header = find_header(ws)
            if not header:
                continue

            c_code = col_index(header, "kazanım kod")
            c_grade = col_index(header, "sınıf düzey")
            c_unit_no = col_index(header, "ünite", "no") if col_index(header, "ünite", "no & ad") is None else None
            c_unit_combined = col_index(header, "no & ad")
            c_unit_name = col_index(header, "ünite", "ad")
            c_subarea = col_index(header, "alt alan")
            c_desc = col_index(header, "öğrenme çıktı")

            sheet_grade = grade_from_sheet(sheet)
            # Okul Öncesi: sayfalar alt alan (Türkçe/Matematik...), sınıf okul öncesi
            preschool = "öncesi" in fname.translate(TR_LOWER).lower()

            order_counter = 0
            for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                if r is None:
                    continue
                code = r[c_code] if c_code is not None and c_code < len(r) else None
                code = str(code).strip().rstrip(".") if code else ""
                if not code or "kod" in code.lower():
                    continue
                desc = r[c_desc] if c_desc is not None and c_desc < len(r) else None
                desc = str(desc).strip() if desc else ""
                if not desc:
                    continue

                # grade
                if c_grade is not None and c_grade < len(r) and r[c_grade]:
                    gcode = grade_from_sheet(r[c_grade]) or str(r[c_grade]).strip()
                elif preschool:
                    gcode = "okul-oncesi"
                else:
                    gcode = sheet_grade
                if not gcode:
                    gcode = "0"

                # unit
                if c_unit_combined is not None and c_unit_combined < len(r):
                    unit = r[c_unit_combined]
                elif c_unit_name is not None and c_unit_name < len(r):
                    unit = r[c_unit_name]
                elif c_subarea is not None and c_subarea < len(r):
                    unit = r[c_subarea]
                else:
                    unit = None
                unit = str(unit).strip() if unit else (sheet if preschool else "Genel")

                order_counter += 1
                unit_order = first_int(r[c_unit_no]) if (c_unit_no is not None and c_unit_no < len(r)) else None
                if unit_order is None:
                    unit_order = order_counter

                if code in seen_codes:
                    skipped += 1
                    continue
                seen_codes.add(code)

                rows.append({
                    "gradeCode": str(gcode),
                    "subjectCode": subj_code,
                    "subjectName": subj_name,
                    "unitName": unit,
                    "unitOrder": unit_order,
                    "outcomeCode": code,
                    "outcomeDescription": desc,
                })
        wb.close()

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1)

    subjects = sorted({(r["subjectCode"], r["subjectName"]) for r in rows})
    print(f"Yazıldı: {args.out}")
    print(f"Kazanım: {len(rows)} (dedupe ile atlanan: {skipped})")
    print(f"Ders sayısı: {len(subjects)}")
    for c, n in subjects:
        print(f"  - {c:24s} {n}")


if __name__ == "__main__":
    main()
